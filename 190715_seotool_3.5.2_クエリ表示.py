#サードパーティーモジュールをインポート
import webbrowser,bs4,requests,os,random
from datetime import datetime, timedelta
import openpyxl as px
import tkinter.filedialog as tkfd
import tkinter
from openpyxl.styles.fonts import Font
from time import sleep

time=datetime.now()

#初期画面の表示
print(u"検索ワードを入力")
x=input()
print(u"取得順位数を半角入力")
r=input()
ranking=int(r)
print()
print("解析中…")
print()

if ranking%10 == 0 :
    y=ranking/10-1
else:
    y=(ranking-ranking%10)/10

top20_URLs=[]
contents_date=[]
headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36"}


for i in range(int(y)+1):
    #Google検索ページから情報を取得
    res=requests.get("https://www.google.co.jp/search?q="+x.replace(" ","+")+"&start="+str(i*10),headers=headers)
    #res.raise_for_status()
    soup=bs4.BeautifulSoup(res.content,"html.parser")
    top_URLs1=soup.select(".r > a")
    top20_URLs=top20_URLs+top_URLs1

    contents_date1=soup.select("span.f")
    contents_date=contents_date+contents_date1

    if i==0:
        suggests=soup.find_all("p")
        resultstatus=soup.select("#resultStats")

#Excel全体フォーマットを自動生成
wb=px.Workbook()
ws=wb.active
    
ws.title=str(x)

fill=px.styles.PatternFill(patternType='solid',fgColor='00CCFF',bgColor='00CCFF')

ws.cell(row=1,column=2).value=u"出力日時"
ws.cell(row=1,column=3).value=time.strftime('%Y/%m/%d %H:%M:%S')
ws.cell(row=1,column=5).value=str(x)
ws.cell(row=3,column=2).value=u"掲載順位"


ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 9
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 50


#繰り返し：各サイトの情報を解析
i=0
for i in range(len(top20_URLs)):
    sleep(random.uniform(0.3,3))
    try:     
        msr=ws.max_row+1
        
        ws.cell(row=msr,column=2).value=i+1
        ws.cell(row=msr,column=3).value=u"タイトル(タイトルクリックでサイトにジャンプ)"
        ws.cell(row=msr,column=3).fill=fill
        ws.cell(row=msr,column=4).value=u"h2見出し"
        ws.cell(row=msr,column=4).fill=fill
        ws.cell(row=msr,column=5).value=u"h3見出し"
        ws.cell(row=msr,column=5).fill=fill
      
        res_contents=requests.get(top20_URLs[i].get("href"),headers=headers)
        #res_contents.raise_for_status()
        contents_soup=bs4.BeautifulSoup(res_contents.content,"html.parser")
        contents_title=contents_soup.select("title")
        h2_elems=contents_soup.select("h2")
        h3_elems=contents_soup.select("h3")

        #1サイト分の情報をExcelに入力
        ws.cell(row=msr+1,column=3).value=top20_URLs[i].getText()[:top20_URLs[i].getText().find("http")]
           
        ws.cell(row=msr+1,column=3).hyperlink =top20_URLs[i].get("href")
        font=Font(color='FF0000FF',underline="single")
        ws.cell(row=msr+1,column=3).font=font

        ws.cell(row=msr+2,column=3).value=top20_URLs[i].get("href")

        """日付取得 コメントアウト中
            date=contents_date[i].getText()[:contents_date[i].getText().find("...")]
        if len(date)<=11:
            ws.cell(row=msr+3,column=3).value=date"""
                   
        j=0
        for j in range(len(h2_elems)-1):
            ws.cell(row=msr+j+1,column=4).value=h2_elems[j].getText()
        for j in range(len(h3_elems)-1):
            ws.cell(row=msr+j+1,column=5).value=h3_elems[j].getText()

        print(str(i+1)+"/"+r+"サイトの解析に成功")
    except:
        print(str(i+1)+"/"+r+"サイトの解析に失敗")

#検索ヒット数を入力
ws.cell(row=1,column=4).value=resultstatus[0].getText()

#サジェストキーワードの取得
ws.cell(row=4,column=1).value=u"サジェスト"
ws.cell(row=4,column=1).fill=fill
i=1
for i in range(len(suggests)-1):
  ws.cell(row=5+i,column=1).value=suggests[i].getText()
print()
print("保存先を選択してください")

#ファイルを保存
tk=tkinter.Tk()
wb.save(tkfd.asksaveasfilename(filetypes=[(u"Excel ブック",(".xlsx"))],initialfile=str(time.strftime('%Y%m%d_'))+x+u"_上位"+r+"サイト解析.xlsx",initialdir="C:/Users"))
tk.withdraw()
